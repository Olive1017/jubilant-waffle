import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

# 全局变量，存储Excel文件和工作簿
_excel_file = None
_workbook = None
_worksheet = None
_global_number = 1


def init_excel_file(file_path):
    """初始化Excel文件，创建表头"""
    global _excel_file, _workbook, _worksheet, _global_number

    _excel_file = Path(file_path)
    _global_number = 1

    # 创建工作簿
    _workbook = openpyxl.Workbook()
    _worksheet = _workbook.active
    _worksheet.title = "合同信息"

    # 添加表头
    headers = [
        '序号', '经办时间', '经办机构', '经办人员',
        '合同电子编号', '合同名称', '处理人', '节点名称',
        '所在部门', '处理意见', '接收时间', '审批时间'
    ]
    _worksheet.append(headers)

    # 设置表头样式（加粗 + 灰色背景）
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    for cell in _worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 调整列宽
    _worksheet.column_dimensions['A'].width = 6
    _worksheet.column_dimensions['B'].width = 20
    _worksheet.column_dimensions['C'].width = 30
    _worksheet.column_dimensions['D'].width = 12
    _worksheet.column_dimensions['E'].width = 32
    _worksheet.column_dimensions['F'].width = 40
    _worksheet.column_dimensions['G'].width = 12
    _worksheet.column_dimensions['H'].width = 20
    _worksheet.column_dimensions['I'].width = 20
    _worksheet.column_dimensions['J'].width = 60
    _worksheet.column_dimensions['K'].width = 20
    _worksheet.column_dimensions['L'].width = 20

    # 保存文件
    _workbook.save(str(_excel_file))
    print(f"Excel文件已创建: {_excel_file}")


def add_contract_data(contract_data):
    """实时追加合同数据到Excel"""
    global _global_number, _workbook, _worksheet

    if _workbook is None:
        print("警告：Excel文件未初始化，无法写入数据")
        return

    basic_info = {
        '经办时间': contract_data.get('经办时间', ''),
        '经办机构': contract_data.get('经办机构', ''),
        '经办人员': contract_data.get('经办人员', ''),
        '合同电子编号': contract_data.get('合同电子编号', ''),
        '合同名称': contract_data.get('合同名称', '')
    }

    approval_records = contract_data.get('审批记录', [])

    if approval_records:
        for record in approval_records:
            row_data = [
                _global_number,
                basic_info['经办时间'],
                basic_info['经办机构'],
                basic_info['经办人员'],
                basic_info['合同电子编号'],
                basic_info['合同名称'],
                record.get('处理人', ''),
                record.get('节点名称', ''),
                record.get('所在部门', ''),
                record.get('处理意见', ''),
                record.get('接收时间', ''),
                record.get('审批时间', '')
            ]
            _worksheet.append(row_data)

            # 设置样式
            for idx, cell in enumerate(_worksheet[_worksheet.max_row], start=1):
                if idx not in [3, 6, 10]:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            _global_number += 1
    else:
        row_data = [
            _global_number,
            basic_info['经办时间'],
            basic_info['经办机构'],
            basic_info['经办人员'],
            basic_info['合同电子编号'],
            basic_info['合同名称'],
            '', '', '', '', '', ''
        ]
        _worksheet.append(row_data)

        # 设置样式
        for idx, cell in enumerate(_worksheet[_worksheet.max_row], start=1):
            if idx not in [3, 6, 10]:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        _global_number += 1

    # 保存文件
    _workbook.save(str(_excel_file))


def save_and_close():
    """保存并关闭Excel文件"""
    global _excel_file, _workbook, _worksheet

    if _workbook is not None:
        _workbook.save(str(_excel_file))
        _workbook.close()
        print(f"Excel文件已保存: {_excel_file}")

        # 重置全局变量
        _excel_file = None
        _workbook = None
        _worksheet = None
